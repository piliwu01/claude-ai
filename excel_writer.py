import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation

# 藍色，較淺 80%（Office 主題 Accent 1, tint 0.8）
_JIAN_COLOR = Color(theme=4, tint=0.8)
JIAN_FILL = PatternFill(fill_type='solid', fgColor=_JIAN_COLOR, bgColor=_JIAN_COLOR)
HEADER_ROW = 2   # 總表第2列為欄位標題
DATA_START = 3   # 資料從第3列開始


def find_next_row(ws):
    """找總表第一個空白資料列"""
    row = DATA_START
    while True:
        if ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=2).value is None:
            return row
        row += 1


def check_duplicate(ws, teacher, days):
    """回傳已存在於總表的 (教師, 星期) 組合"""
    existing = set()
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if row[1] == teacher and row[5] in days:
            existing.add(row[5])
    return existing


def build_subject_code_map(wb):
    """
    從 工作表1 欄 A 建立科別名稱 → 完整代碼的對照表
    例如：'國文' → '_01國文'，'英語' → '_02英語'
    """
    ws = wb['工作表1']
    code_map = {}
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).startswith('_') and len(str(val)) > 3:
            code = str(val).strip()      # '_01國文'
            name = code[3:].strip()      # '國文'
            if name:
                code_map[name] = code
    return code_map


def _add_validations(ws, start_row, end_row):
    """為新寫入的列加上與現有欄位一致的下拉驗證"""
    # 欄 A（請假科別）和 欄 C（代課科別）：命名範圍「科別」
    dv_subj = DataValidation(
        type="list", formula1="科別",
        allow_blank=True, showErrorMessage=False)
    dv_subj.sqref = f"A{start_row}:A{end_row} C{start_row}:C{end_row}"
    ws.add_data_validation(dv_subj)

    # 欄 B（請假教師）：INDIRECT($A<row>) 依請假科別連動
    dv_b = DataValidation(
        type="list", formula1=f"INDIRECT($A{start_row})",
        allow_blank=True, showErrorMessage=False)
    dv_b.sqref = f"B{start_row}:B{end_row}"
    ws.add_data_validation(dv_b)

    # 欄 D（代課教師）：INDIRECT($C<row>) 依代課科別連動
    dv_d = DataValidation(
        type="list", formula1=f"INDIRECT($C{start_row})",
        allow_blank=True, showErrorMessage=False)
    dv_d.sqref = f"D{start_row}:D{end_row}"
    ws.add_data_validation(dv_d)

    # 欄 F（星期）：工作表1 X 欄
    dv_f = DataValidation(
        type="list", formula1="工作表1!$X$2:$X$7",
        allow_blank=True, showErrorMessage=False)
    dv_f.sqref = f"F{start_row}:F{end_row}"
    ws.add_data_validation(dv_f)

    # 欄 G（節次）：工作表1 Y 欄
    dv_g = DataValidation(
        type="list", formula1="工作表1!$Y$2:$Y$10",
        allow_blank=True, showErrorMessage=False)
    dv_g.sqref = f"G{start_row}:G{end_row}"
    ws.add_data_validation(dv_g)

    # 欄 H（科目）：工作表1 AA 欄清單
    dv_h = DataValidation(
        type="list", formula1="工作表1!$AA$2:$AA$33",
        allow_blank=True, showErrorMessage=False)
    dv_h.sqref = f"H{start_row}:H{end_row}"
    ws.add_data_validation(dv_h)

    # 欄 J（假別）：工作表1 Z 欄清單
    dv_j = DataValidation(
        type="list", formula1="工作表1!$Z$2:$Z$8",
        allow_blank=True, showErrorMessage=False)
    dv_j.sqref = f"J{start_row}:J{end_row}"
    ws.add_data_validation(dv_j)


def write_to_zongbiao(excel_path, teacher, subject, records):
    """
    records: list of {'星期', '節次', '科目', '班級', '兼課'}
    回傳寫入筆數
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['總表']

    # 科別名稱 → '_01國文' 格式
    code_map = build_subject_code_map(wb)
    subject_code = code_map.get(subject, subject)  # 找不到就用原值

    start_row = find_next_row(ws)

    for i, r in enumerate(records):
        row = start_row + i
        ws.cell(row=row, column=1).value = subject_code   # 科別（請假）帶代碼
        ws.cell(row=row, column=2).value = teacher        # 請假教師
        ws.cell(row=row, column=3).value = None           # 科別（代課）
        ws.cell(row=row, column=4).value = None           # 代課教師
        ws.cell(row=row, column=5).value = None           # 日期
        ws.cell(row=row, column=6).value = r['星期']      # 星期
        ws.cell(row=row, column=7).value = r['節次']      # 節次
        ws.cell(row=row, column=8).value = r['科目']      # 科目
        班級 = r['班級']
        ws.cell(row=row, column=9).value = int(班級) if str(班級).isdigit() else 班級  # 班級
        ws.cell(row=row, column=10).value = '公假'         # 假別（預設公假）

        if r['兼課']:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = JIAN_FILL

    end_row = start_row + len(records) - 1
    _add_validations(ws, start_row, end_row)

    wb.save(excel_path)
    return len(records)
