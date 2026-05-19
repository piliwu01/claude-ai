import warnings
warnings.filterwarnings("ignore")

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
import os
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins

# ── GUI 色彩 ──────────────────────────────────────────────────────────────────
BG       = "#f0f4f8"
CARD_BG  = "#ffffff"
BTN_BLUE = "#2563eb"
BTN_GRN  = "#16a34a"
TITLE_FG = "#1e293b"
SUB_FG   = "#64748b"
LOG_BG   = "#1e293b"

# ── Excel 樣式 ────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1e3a5f")   # 深藍（關鍵欄）
HDR_FILL2  = PatternFill("solid", fgColor="ffffff")   # 白（其他欄）
ROW_FILL   = PatternFill("solid", fgColor="dce6f1")   # 淺藍（資料列）

THIN   = Side(style="thin",   color="b0b8c4")
MEDIUM = Side(style="medium", color="000000")
HDR_BORDER = Border(left=THIN, right=THIN, top=MEDIUM, bottom=MEDIUM)

NUM_COLS = 8
DISPLAY_COLS = ["請假教師", "代課教師", "日期", "星期", "節次", "科目", "班級", "假別"]
COL_WIDTHS   = [13, 13, 10, 4.5, 4.5, 8, 7, 7]

NOTICE_CONFIGS = {
    "請假教師": {
        "title":    "請假老師課務移交通知單",
        "subtitle": None,
        "group_col": "請假教師",
        "sheet":    "請假教師",
    },
    "代課教師": {
        "title":    "代課教師通知單",
        "subtitle": "因為任課老師請假，特請您協助代課。",
        "group_col": "代課教師",
        "sheet":    "代課教師",
    },
    "班級": {
        "title":    "班級代課通知單",
        "subtitle": None,
        "group_col": "班級",
        "sheet":    "班級通知",
    },
}

COL_MAP = {
    "請假教師": "請假教師", "代課教師": "代課教師",
    "日期": "日期_顯示", "星期": "星期", "節次": "節次",
    "科目": "科目", "班級": "班級", "假別": "假別",
}


def load_data(filepath, start_date, end_date):
    df = pd.read_excel(filepath, sheet_name="總表", header=None, engine="openpyxl")
    df = df.iloc[2:].reset_index(drop=True)

    base_cols = ["科別_假", "請假教師", "科別_代", "代課教師", "日期", "星期", "節次", "科目", "班級", "假別"]
    n = df.shape[1]
    if n > 10:
        base_cols.append("取消")
    extra = [f"_e{i}" for i in range(n - len(base_cols))]
    df.columns = base_cols + extra

    df = df.dropna(subset=["請假教師"])
    if "取消" in df.columns:
        df = df[df["取消"].astype(str).str.strip() != "取消"]

    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    df = df.dropna(subset=["日期"])
    df = df[(df["日期"].dt.date >= start_date) & (df["日期"].dt.date <= end_date)]

    df["日期_顯示"] = df["日期"].apply(lambda x: f"{x.month}月{x.day}日")
    df["節次"] = pd.to_numeric(df["節次"], errors="coerce").fillna(0).astype(int).astype(str)
    df = df.sort_values(["日期", "節次"]).reset_index(drop=True)
    return df


def write_notice_sheet(ws, df, notice_type, print_date_str):
    cfg       = NOTICE_CONFIGS[notice_type]
    group_col = cfg["group_col"]
    bold_idx  = DISPLAY_COLS.index(group_col)  # 0-based

    # 清除舊資料與分頁符號
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    ws.row_breaks.brk.clear()

    # 欄寬
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 列印設定：B5 橫式
    ws.page_setup.paperSize        = 13         # B5
    ws.page_setup.orientation      = "landscape"
    ws.page_setup.fitToPage        = False
    ws.page_setup.scale             = 100
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.2, footer=0.2
    )
    ws.print_options.horizontalCentered = True

    groups      = list(df.groupby(group_col, sort=True))
    current_row = 1

    for g_idx, (key, gdf) in enumerate(groups):
        gdf = gdf.sort_values(["日期", "節次"])
        start_of_group = current_row

        # ── 標題列 ──────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=NUM_COLS)
        c = ws.cell(row=current_row, column=1, value=cfg["title"])
        c.font      = Font(name="微軟正黑體", bold=True, size=18)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 36
        current_row += 1

        # ── 副標列（代課教師才有）──────────────────────
        if cfg["subtitle"]:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=NUM_COLS)
            c = ws.cell(row=current_row, column=1, value=cfg["subtitle"])
            c.font      = Font(name="微軟正黑體", size=12, color="475569")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 22
            current_row += 1

        # ── 表頭列 ──────────────────────────────────────
        for ci, col_name in enumerate(DISPLAY_COLS, 1):
            is_key = (ci - 1 == bold_idx)
            c = ws.cell(row=current_row, column=ci, value=col_name)
            c.font      = Font(name="微軟正黑體", bold=True,
                               size=14 if is_key else 12,
                               color="FFFFFF" if is_key else "000000")
            c.fill      = HDR_FILL if is_key else HDR_FILL2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = HDR_BORDER
        ws.row_dimensions[current_row].height = 26
        current_row += 1

        # ── 資料列 ──────────────────────────────────────
        for r_idx, (_, row) in enumerate(gdf.iterrows()):
            fill = ROW_FILL if r_idx % 2 == 0 else PatternFill("solid", fgColor="ffffff")
            for ci, col_name in enumerate(DISPLAY_COLS, 1):
                val = str(row.get(COL_MAP[col_name], ""))
                c = ws.cell(row=current_row, column=ci, value=val)
                c.font      = Font(name="微軟正黑體", bold=True, size=12)
                c.fill      = fill
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 22
            current_row += 1

        # ── 頁尾列 ──────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=NUM_COLS)
        c = ws.cell(row=current_row, column=1,
                    value=f"共 {len(gdf)} 節　　列印日期：{print_date_str}")
        c.font      = Font(name="微軟正黑體", size=9, color="64748b")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # ── 分頁符號（最後一組不加）──────────────────────
        if g_idx < len(groups) - 1:
            ws.row_breaks.append(Break(id=current_row - 1))

    # 列印範圍
    ws.print_area = f"A1:{get_column_letter(NUM_COLS)}{current_row - 1}"
    return len(groups)


# ── GUI ───────────────────────────────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        root.title("代課通知單列印工具 v2.0")
        root.configure(bg=BG)
        root.resizable(False, False)
        root.geometry("520x660")
        self._build()
        self._set_this_week()

    def _lf(self, parent, text):
        return tk.LabelFrame(parent, text=f"  {text}  ",
                             font=("微軟正黑體", 10, "bold"),
                             bg=CARD_BG, fg="#334155",
                             relief="groove", bd=1, padx=10, pady=8)

    def _btn(self, parent, text, color, cmd, **kw):
        return tk.Button(parent, text=text, bg=color, fg="white",
                         font=("微軟正黑體", 10, "bold"), relief="flat",
                         cursor="hand2", command=cmd, **kw)

    def _grey_btn(self, parent, text, cmd):
        return tk.Button(parent, text=text, bg="#e2e8f0",
                         font=("微軟正黑體", 9), relief="flat",
                         padx=8, cursor="hand2", command=cmd)

    def _build(self):
        pk = dict(padx=14, pady=5, fill="x")

        h = tk.Frame(self.root, bg=BG)
        h.pack(fill="x", padx=14, pady=(10, 4))
        tk.Label(h, text="代課通知單列印工具", font=("微軟正黑體", 18, "bold"),
                 bg=BG, fg=TITLE_FG).pack(anchor="w")
        tk.Label(h, text="自動依老師 / 班級分頁，在 Excel 直接列印",
                 font=("微軟正黑體", 10), bg=BG, fg=SUB_FG).pack(anchor="w")

        # 來源檔案
        c1 = self._lf(self.root, "📁 來源檔案（含總表）")
        c1.pack(**pk)
        r = tk.Frame(c1, bg=CARD_BG); r.pack(fill="x")
        self.file_var = tk.StringVar()
        tk.Entry(r, textvariable=self.file_var, font=("微軟正黑體", 9),
                 relief="solid", bd=1, width=44).pack(side="left", ipady=4)
        self._btn(r, "選擇", BTN_BLUE, self._pick_file, padx=8).pack(side="left", padx=(6, 0))

        # 週次
        c2 = self._lf(self.root, "📅 列印週次")
        c2.pack(**pk)
        r2 = tk.Frame(c2, bg=CARD_BG); r2.pack(fill="x")
        tk.Label(r2, text="起始", font=("微軟正黑體", 10), bg=CARD_BG).pack(side="left")
        self.start_var = tk.StringVar()
        tk.Entry(r2, textvariable=self.start_var, font=("微軟正黑體", 10),
                 relief="solid", bd=1, width=12).pack(side="left", padx=4, ipady=4)
        tk.Label(r2, text="～", font=("微軟正黑體", 10), bg=CARD_BG).pack(side="left")
        self.end_var = tk.StringVar()
        tk.Entry(r2, textvariable=self.end_var, font=("微軟正黑體", 10),
                 relief="solid", bd=1, width=12).pack(side="left", padx=4, ipady=4)
        self._grey_btn(r2, "本週", self._set_this_week).pack(side="left", padx=(8, 2))
        self._grey_btn(r2, "上週", self._set_last_week).pack(side="left", padx=2)
        tk.Label(c2, text="格式：YYYY-MM-DD", font=("微軟正黑體", 8),
                 bg=CARD_BG, fg=SUB_FG).pack(anchor="w")

        # 輸出資料夾
        c3 = self._lf(self.root, "📂 輸出資料夾")
        c3.pack(**pk)
        r3 = tk.Frame(c3, bg=CARD_BG); r3.pack(fill="x")
        self.out_var = tk.StringVar()
        tk.Entry(r3, textvariable=self.out_var, font=("微軟正黑體", 9),
                 relief="solid", bd=1, width=44).pack(side="left", ipady=4)
        self._btn(r3, "選擇", BTN_BLUE, self._pick_out, padx=8).pack(side="left", padx=(6, 0))

        # 輸出項目
        c4 = self._lf(self.root, "🖨️ 輸出項目")
        c4.pack(**pk)
        self.chk = {}
        labels = {"請假教師": "請假教師通知單", "代課教師": "代課教師通知單", "班級": "班級代課通知單"}
        for key, disp in labels.items():
            v = tk.BooleanVar(value=True)
            self.chk[key] = v
            tk.Checkbutton(c4, text=disp, variable=v,
                           font=("微軟正黑體", 10), bg=CARD_BG, fg=TITLE_FG,
                           activebackground=CARD_BG).pack(anchor="w")

        self.run_btn = self._btn(self.root, "更新 Excel 並開啟", BTN_GRN, self._run,
                                  padx=30, pady=10)
        self.run_btn.configure(font=("微軟正黑體", 13, "bold"))
        self.run_btn.pack(pady=8)

        self.prog = ttk.Progressbar(self.root, orient="horizontal",
                                    length=490, mode="determinate")
        self.prog.pack(padx=14, pady=(0, 6))

        lf = tk.Frame(self.root, bg=BG)
        lf.pack(fill="both", expand=True, padx=14, pady=(0, 10))
        self.log_box = scrolledtext.ScrolledText(lf, height=8,
                                                  font=("Consolas", 9),
                                                  bg=LOG_BG, fg="#94a3b8",
                                                  relief="flat", wrap="word",
                                                  state="disabled")
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_config("ok",   foreground="#4ade80")
        self.log_box.tag_config("warn", foreground="#f87171")
        self.log_box.tag_config("sep",  foreground="#475569")

    def _pick_file(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")])
        if p:
            self.file_var.set(p)
            if not self.out_var.get():
                self.out_var.set(os.path.dirname(p))

    def _pick_out(self):
        p = filedialog.askdirectory()
        if p:
            self.out_var.set(p)

    def _set_this_week(self):
        today = datetime.now()
        mon = today - timedelta(days=today.weekday())
        fri = mon + timedelta(days=4)
        self.start_var.set(mon.strftime("%Y-%m-%d"))
        self.end_var.set(fri.strftime("%Y-%m-%d"))

    def _set_last_week(self):
        today = datetime.now()
        mon = today - timedelta(days=today.weekday() + 7)
        fri = mon + timedelta(days=4)
        self.start_var.set(mon.strftime("%Y-%m-%d"))
        self.end_var.set(fri.strftime("%Y-%m-%d"))

    def log(self, msg, tag=""):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _run(self):
        fp  = self.file_var.get().strip()
        out = self.out_var.get().strip()
        s   = self.start_var.get().strip()
        e   = self.end_var.get().strip()

        if not fp or not os.path.exists(fp):
            self.log("❌ 請選擇有效的 Excel 檔案", "warn"); return
        if not out or not os.path.isdir(out):
            self.log("❌ 請選擇有效的輸出資料夾", "warn"); return
        selected = [k for k, v in self.chk.items() if v.get()]
        if not selected:
            self.log("❌ 請至少勾選一種通知單", "warn"); return
        try:
            sd = datetime.strptime(s, "%Y-%m-%d").date()
            ed = datetime.strptime(e, "%Y-%m-%d").date()
        except ValueError:
            self.log("❌ 日期格式錯誤，請用 YYYY-MM-DD", "warn"); return

        self.run_btn.configure(state="disabled")
        self.prog["value"] = 0
        threading.Thread(
            target=self._worker, args=(fp, out, sd, ed, selected), daemon=True
        ).start()

    def _worker(self, fp, out_dir, sd, ed, selected):
        try:
            self.log("=== 開始處理 ===", "sep")
            self.log(f"📄 讀取：{os.path.basename(fp)}")
            df = load_data(fp, sd, ed)
            self.log(f"📄 篩選到 {len(df)} 筆（{sd} ～ {ed}）")

            if df.empty:
                self.log("⚠️ 選取週次無任何資料，請確認日期範圍", "warn")
                return

            # 建立新活頁簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 移除預設空白頁

            suffix     = f"{sd.strftime('%m%d')}-{ed.strftime('%m%d')}"
            print_date = datetime.now().strftime("%Y/%m/%d")
            out_path   = os.path.join(out_dir, f"代課通知單_{suffix}.xlsx")

            for i, ntype in enumerate(selected):
                sheet_name = NOTICE_CONFIGS[ntype]["sheet"]
                ws = wb.create_sheet(title=sheet_name)
                count = write_notice_sheet(ws, df, ntype, print_date)
                self.log(f"✅ {ntype}：{count} 人/班 → 工作表「{sheet_name}」", "ok")
                self.prog["value"] = (i + 1) / len(selected) * 100
                self.root.update_idletasks()

            wb.save(out_path)
            self.log("=== 完成 ===", "sep")
            self.log(f"✅ 已儲存：{os.path.basename(out_path)}", "ok")

            if messagebox.askyesno("完成", f"已產生 代課通知單_{suffix}.xlsx\n要開啟 Excel 嗎？"):
                os.startfile(out_path)

        except Exception as ex:
            import traceback
            self.log(f"❌ 錯誤：{ex}", "warn")
            self.log(traceback.format_exc(), "warn")
        finally:
            self.run_btn.configure(state="normal")


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
