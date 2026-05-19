import warnings
warnings.filterwarnings("ignore")

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext
import threading
import os
import sys
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 色彩與字型 ───────────────────────────────────────────
BG       = "#f0f4f8"
CARD     = "#ffffff"
BLUE_BTN = "#2563eb"
GREEN_BTN= "#16a34a"
TITLE_FG = "#1e293b"
SUB_FG   = "#64748b"
LOG_BG   = "#1e293b"
LOG_FG   = "#94a3b8"
OK_FG    = "#4ade80"
ERR_FG   = "#f87171"
SEP_FG   = "#475569"
FONT     = "微軟正黑體"
MONO     = "Consolas"

# ── 找程式自身目錄（用於讀 assets）─────────────────────
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ══════════════════════════════════════════════════════════
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("師課表第8節節數計算器 v1.0")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.root.minsize(680, 720)
        self.root.geometry("800x860")

        self.pdf_path     = tk.StringVar()
        self.output_dir   = tk.StringVar()
        self.output_name  = tk.StringVar(value="114-2_一二年級第8節.xlsx")
        # 星期勾選（預設一、二）
        self.day_vars = {
            '一': tk.BooleanVar(value=True),
            '二': tk.BooleanVar(value=True),
            '三': tk.BooleanVar(value=False),
            '四': tk.BooleanVar(value=False),
            '五': tk.BooleanVar(value=False),
        }
        # 年級勾選（預設一、二年級）
        self.grade_vars = {
            '一': tk.BooleanVar(value=True),
            '二': tk.BooleanVar(value=True),
            '三': tk.BooleanVar(value=False),
        }
        # 週別
        self.week_mode = tk.StringVar(value='單')

        self._build_ui()

    # ── 建立 UI ──────────────────────────────────────────
    def _card(self, parent, title):
        f = tk.LabelFrame(parent, text=f"  {title}  ",
                          font=(FONT, 10, "bold"),
                          bg=CARD, fg="#334155",
                          relief="groove", bd=1,
                          padx=12, pady=10)
        f.pack(fill="x", padx=16, pady=6)
        return f

    def _build_ui(self):
        # 大標題
        hdr = tk.Frame(self.root, bg=BG)
        hdr.pack(fill="x", padx=16, pady=(16, 4))
        tk.Label(hdr, text="師課表第 8 節節數計算器",
                 font=(FONT, 18, "bold"), fg=TITLE_FG, bg=BG).pack(anchor="w")
        tk.Label(hdr, text="從師課表 PDF 統計指定星期與年級的第 8 節節數，並填入共同扣除",
                 font=(FONT, 10), fg=SUB_FG, bg=BG).pack(anchor="w")

        # Card 1：PDF
        c1 = self._card(self.root, "① 師課表 PDF")
        r1 = tk.Frame(c1, bg=CARD)
        r1.pack(fill="x")
        tk.Button(r1, text="選擇檔案", bg=BLUE_BTN, fg="white",
                  font=(FONT, 10, "bold"), relief="flat", padx=10, cursor="hand2",
                  command=self._pick_pdf).pack(side="right")
        tk.Entry(r1, textvariable=self.pdf_path, font=(FONT, 10),
                 relief="solid", bd=1).pack(side="left", fill="x", expand=True, ipady=4, padx=(0, 8))

        # Card 2：輸出
        c2 = self._card(self.root, "② 輸出位置")
        r2a = tk.Frame(c2, bg=CARD)
        r2a.pack(fill="x", pady=(0, 6))
        tk.Button(r2a, text="選擇資料夾", bg=BLUE_BTN, fg="white",
                  font=(FONT, 10, "bold"), relief="flat", padx=10, cursor="hand2",
                  command=self._pick_dir).pack(side="right")
        tk.Entry(r2a, textvariable=self.output_dir, font=(FONT, 10),
                 relief="solid", bd=1).pack(side="left", fill="x", expand=True, ipady=4, padx=(0, 8))
        r2b = tk.Frame(c2, bg=CARD)
        r2b.pack(fill="x")
        tk.Label(r2b, text="檔名：", font=(FONT, 10), bg=CARD, fg=TITLE_FG).pack(side="left")
        tk.Entry(r2b, textvariable=self.output_name, font=(FONT, 10),
                 relief="solid", bd=1).pack(side="left", fill="x", expand=True, ipady=4)

        # Card 3：設定
        c3 = self._card(self.root, "③ 計算設定")
        c3.columnconfigure(1, weight=1)

        # 統計星期
        tk.Label(c3, text="統計星期：", font=(FONT, 10, "bold"),
                 bg=CARD, fg=TITLE_FG).grid(row=0, column=0, sticky="w", pady=5)
        df = tk.Frame(c3, bg=CARD)
        df.grid(row=0, column=1, sticky="w", pady=5)
        for d, var in self.day_vars.items():
            tk.Checkbutton(df, text=f"星期{d}", variable=var,
                           font=(FONT, 10), bg=CARD, fg=TITLE_FG,
                           activebackground=CARD,
                           selectcolor=CARD).pack(side="left", padx=6)

        # 統計年級
        tk.Label(c3, text="統計年級：", font=(FONT, 10, "bold"),
                 bg=CARD, fg=TITLE_FG).grid(row=1, column=0, sticky="w", pady=5)
        gf = tk.Frame(c3, bg=CARD)
        gf.grid(row=1, column=1, sticky="w", pady=5)
        for g, var in self.grade_vars.items():
            tk.Checkbutton(gf, text=f"{g}年級", variable=var,
                           font=(FONT, 10), bg=CARD, fg=TITLE_FG,
                           activebackground=CARD,
                           selectcolor=CARD).pack(side="left", padx=6)

        # 計算週別
        tk.Label(c3, text="計算週別：", font=(FONT, 10, "bold"),
                 bg=CARD, fg=TITLE_FG).grid(row=2, column=0, sticky="w", pady=5)
        wf = tk.Frame(c3, bg=CARD)
        wf.grid(row=2, column=1, sticky="w", pady=5)
        for label, val in [("單週（奇數週）", "單"), ("雙週（偶數週）", "雙"), ("全部（單＋雙）", "全部")]:
            tk.Radiobutton(wf, text=label, variable=self.week_mode, value=val,
                           font=(FONT, 10), bg=CARD, fg=TITLE_FG,
                           activebackground=CARD,
                           selectcolor=CARD).pack(side="left", padx=8)

        # 執行按鈕
        btn_row = tk.Frame(self.root, bg=BG)
        btn_row.pack(pady=(8, 4))
        tk.Button(btn_row, text="開始計算", bg=GREEN_BTN, fg="white",
                  font=(FONT, 13, "bold"), relief="flat", padx=30, pady=10,
                  cursor="hand2", command=self._run).pack(side="left", padx=6)
        tk.Button(btn_row, text="清除記錄", bg="#e2e8f0",
                  font=(FONT, 9), relief="flat", padx=8,
                  cursor="hand2", command=self._clear_log).pack(side="left", padx=6)

        # 進度條
        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", pady=4, padx=16)

        # 執行記錄（填滿剩餘空間）
        self.log = scrolledtext.ScrolledText(
            self.root, font=(MONO, 9), bg=LOG_BG, fg=LOG_FG,
            relief="flat", state="disabled")
        self.log.pack(fill="both", expand=True, padx=16, pady=(4, 16))
        self.log.tag_config("ok",  foreground=OK_FG)
        self.log.tag_config("err", foreground=ERR_FG)
        self.log.tag_config("sep", foreground=SEP_FG)

    # ── 事件 ─────────────────────────────────────────────
    def _pick_pdf(self):
        p = filedialog.askopenfilename(
            title="選擇師課表 PDF",
            filetypes=[("PDF 檔案", "*.pdf"), ("所有檔案", "*.*")])
        if p:
            self.pdf_path.set(p)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(p))

    def _pick_dir(self):
        p = filedialog.askdirectory(title="選擇輸出資料夾")
        if p:
            self.output_dir.set(p)

    def _log(self, msg, tag=None):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag or "")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _clear_log(self):
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")
        self.progress["value"] = 0

    def _run(self):
        pdf = self.pdf_path.get().strip()
        out_dir = self.output_dir.get().strip()
        filename = self.output_name.get().strip() or "第8節節數.xlsx"

        if not pdf:
            self._log("❌ 請先選擇師課表 PDF", "err"); return
        if not os.path.exists(pdf):
            self._log("❌ PDF 檔案不存在，請重新選擇", "err"); return
        if not out_dir:
            out_dir = os.path.dirname(pdf)
            self.output_dir.set(out_dir)

        selected_days   = [d for d, v in self.day_vars.items()  if v.get()]
        selected_grades = [g for g, v in self.grade_vars.items() if v.get()]
        week_mode = self.week_mode.get()

        if not selected_days:
            self._log("❌ 請至少勾選一個星期", "err"); return
        if not selected_grades:
            self._log("❌ 請至少勾選一個年級", "err"); return

        output_path = os.path.join(out_dir, filename)
        self.progress["value"] = 0
        threading.Thread(
            target=self._calc,
            args=(pdf, output_path, selected_days, selected_grades, week_mode),
            daemon=True
        ).start()

    # ── 核心計算（背景執行緒）────────────────────────────
    def _calc(self, pdf_path, output_path, selected_days, selected_grades, week_mode):
        try:
            import pdfplumber

            DAY_COL    = {'一': 2, '二': 3, '三': 4, '四': 5, '五': 6}
            GRADE_PFX  = {'一': '1', '二': '2', '三': '3'}
            WEEK_LABEL = {'單': '單週（奇數週）', '雙': '雙週（偶數週）', '全部': '全部（單＋雙）'}

            target_cols    = [DAY_COL[d] for d in selected_days]
            target_pfx     = tuple(GRADE_PFX[g] for g in selected_grades)

            # 讀共同扣除對照表
            deduction_map = {}
            deduct_path = os.path.join(SCRIPT_DIR, '第八節鐘點費扣除.xlsx')
            if os.path.exists(deduct_path):
                wb_d = openpyxl.load_workbook(deduct_path)
                for row in wb_d.active.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        deduction_map[str(row[0]).strip()] = int(row[1]) if row[1] else 0
                self._log(f"📋 已載入扣除對照表（{len(deduction_map)} 人）")
            else:
                self._log("⚠️  未找到第八節鐘點費扣除.xlsx，共同扣除欄全部填 0")

            def parse_cell(cell, mode):
                """
                回傳 (班碼, 節數)
                - 無標記：每週有課 → 計算
                - 單週：奇數週有課
                - 雙週：偶數週有課
                - 混合（單+雙）：每週有課，不同班交替
                mode='單'  → 計算「無標記」與「單週」，跳過「純雙週」
                mode='雙'  → 計算「無標記」與「雙週」，跳過「純單週」
                mode='全部'→ 全部計算
                """
                if not cell or not cell.strip():
                    return None, 0
                has_dan    = '單' in cell
                has_shuang = '雙' in cell
                parts = [p.strip() for p in cell.split('\n') if p.strip()]
                if len(parts) < 2:
                    return None, 0

                codes = re.findall(r'(\d{3})', parts[1])

                if not has_dan and not has_shuang:
                    # 正常每週
                    return (codes[0] if codes else None), 1

                if has_dan and has_shuang:
                    # 混合：每週都有課（奇週A班、偶週B班）→ 一律計 1
                    if mode == '單':
                        return (codes[0] if codes else None), 1
                    elif mode == '雙':
                        return (codes[1] if len(codes) > 1 else codes[0] if codes else None), 1
                    else:  # 全部
                        return (codes[0] if codes else None), 1

                if has_dan and not has_shuang:
                    # 純單週
                    if mode == '雙':
                        return None, 0
                    return (codes[0] if codes else None), 1

                # 純雙週
                if mode == '單':
                    return None, 0
                return (codes[0] if codes else None), 1

            # ── 開始讀 PDF ──────────────────────────────
            self._log("=" * 52, "sep")
            self._log(f"📂 {os.path.basename(pdf_path)}")
            self._log(f"📅 統計星期：{'、'.join(f'星期{d}' for d in selected_days)}")
            self._log(f"🏫 統計年級：{'、'.join(f'{g}年級' for g in selected_grades)}")
            self._log(f"📋 計算週別：{WEEK_LABEL[week_mode]}")
            self._log("=" * 52, "sep")

            results = {}

            with pdfplumber.open(pdf_path) as pdf:
                total = len(pdf.pages)
                self.root.after(0, lambda: self.progress.configure(maximum=total))

                for i, page in enumerate(pdf.pages):
                    self.root.after(0, lambda v=i+1: self.progress.configure(value=v))

                    text  = page.extract_text() or ""
                    lines = text.strip().split('\n')
                    teacher_code = teacher_name = None
                    for line in lines:
                        m = re.match(r'^(\d{3})\s+(\S+)', line)
                        if m:
                            teacher_code, teacher_name = m.group(1), m.group(2)
                            break
                    if not teacher_code or teacher_code in results:
                        continue

                    tables = page.extract_tables()
                    row8 = None
                    if tables:
                        for row in tables[0]:
                            if row and row[0] == '8':
                                row8 = row
                                break

                    day_counts  = {}
                    day_classes = {}

                    for day, col in zip(selected_days, target_cols):
                        cell = (row8[col] if row8 and len(row8) > col else None)
                        code, cnt = parse_cell(cell, week_mode)
                        if code and code[0] in target_pfx:
                            day_counts[day]  = cnt
                            day_classes[day] = code
                        else:
                            day_counts[day]  = 0
                            day_classes[day] = ''

                    results[teacher_code] = {
                        'name':        teacher_name,
                        'day_counts':  day_counts,
                        'day_classes': day_classes,
                        'total':       sum(day_counts.values()),
                        'deduct':      deduction_map.get(teacher_name, 0),
                    }

            sorted_results = sorted(results.items(), key=lambda x: x[0])

            # ── 輸出 Excel ──────────────────────────────
            self._log("📊 正在輸出 Excel...", "ok")
            n_days = len(selected_days)

            # 欄位定義（動態）
            # A:課表代號  B:教師姓名  C~C+n-1:各星期班級  C+n~C+2n-1:各星期節數  合計  共同扣除
            col_class_start = 3                         # 各星期班級欄起始
            col_count_start = 3 + n_days               # 各星期節數欄起始
            col_total       = 3 + n_days * 2            # 合計節數欄
            col_deduct      = 3 + n_days * 2 + 1        # 共同扣除欄
            n_cols          = col_deduct                 # 總欄數

            headers = (
                ["課表代號", "教師姓名"] +
                [f"星期{d}第8節班級" for d in selected_days] +
                [f"星期{d}節數"      for d in selected_days] +
                ["合計節數", "共同扣除"]
            )

            HDR_FILL    = PatternFill("solid", fgColor="1E3A5F")
            ALT_FILL    = PatternFill("solid", fgColor="EBF2FF")
            ZERO_FILL   = PatternFill("solid", fgColor="F5F5F5")
            TOTAL_FILL  = PatternFill("solid", fgColor="D4EDDA")
            DEDUCT_FILL = PatternFill("solid", fgColor="FFF3CD")
            SUM_FILL    = PatternFill("solid", fgColor="FFD700")
            thin   = Side(border_style="thin", color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "第8節節數"

            # 大標題
            last_letter = get_column_letter(n_cols)
            ws.merge_cells(f"A1:{last_letter}1")
            days_str  = "、".join(f"星期{d}" for d in selected_days)
            grades_str = "、".join(f"{g}年級" for g in selected_grades)
            ws["A1"] = f"114-2 {days_str}任教{grades_str}之教師第 8 節節數統計（{WEEK_LABEL[week_mode]}）"
            ws["A1"].font = Font(name=FONT, size=13, bold=True, color="1E3A5F")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # 標題列
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=c, value=h)
                cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            ws.row_dimensions[2].height = 22

            # 資料列
            for r, (code, d) in enumerate(sorted_results, 3):
                is_zero  = d['total'] == 0
                base_fill = ZERO_FILL if is_zero else (ALT_FILL if r % 2 == 0 else PatternFill())

                row_data = (
                    [code, d['name']] +
                    [d['day_classes'].get(day, '') for day in selected_days] +
                    [d['day_counts'].get(day, 0)   for day in selected_days] +
                    [d['total'], d['deduct']]
                )
                for c, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = Font(
                        name=FONT, size=10,
                        bold=(c == col_total and not is_zero),
                        color="999999" if is_zero else "000000")
                    cell.border = border
                    if c == col_total and not is_zero:
                        cell.fill = TOTAL_FILL
                    elif c == col_deduct:
                        cell.fill = DEDUCT_FILL
                    else:
                        cell.fill = base_fill
                    cell.alignment = Alignment(
                        horizontal="left" if c == 2 else "center",
                        vertical="center")
                ws.row_dimensions[r].height = 18

            # 合計列
            last_row = len(sorted_results) + 3
            for c in range(1, n_cols + 1):
                val = ''
                if c == 1:
                    val = '合計'
                elif col_count_start <= c < col_total:
                    day = selected_days[c - col_count_start]
                    val = sum(d['day_counts'].get(day, 0) for _, d in sorted_results)
                elif c == col_total:
                    val = sum(d['total']  for _, d in sorted_results)
                elif c == col_deduct:
                    val = sum(d['deduct'] for _, d in sorted_results)
                cell = ws.cell(row=last_row, column=c, value=val)
                cell.font  = Font(name=FONT, size=10, bold=True)
                cell.fill  = SUM_FILL
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[last_row].height = 20

            # 欄寬
            for c in range(1, n_cols + 1):
                letter = get_column_letter(c)
                if c <= 2:
                    ws.column_dimensions[letter].width = 12
                elif c < col_count_start:   # 班級欄
                    ws.column_dimensions[letter].width = 16
                else:
                    ws.column_dimensions[letter].width = 12

            wb.save(output_path)

            non_zero   = sum(1 for _, d in sorted_results if d['total'] > 0)
            total_sum  = sum(d['total']  for _, d in sorted_results)
            deduct_sum = sum(d['deduct'] for _, d in sorted_results)

            self._log(f"✅ 完成！共 {len(sorted_results)} 位教師", "ok")
            self._log(f"   有第 8 節者：{non_zero} 位，合計 {total_sum} 節", "ok")
            self._log(f"   共同扣除合計：{deduct_sum} 節", "ok")
            self._log(f"📁 已儲存：{output_path}", "ok")
            self.root.after(0, lambda: self.progress.configure(value=self.progress["maximum"]))

        except Exception as e:
            import traceback
            self._log(f"❌ 錯誤：{e}", "err")
            self._log(traceback.format_exc(), "err")


# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
