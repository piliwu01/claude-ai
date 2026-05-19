"""執行一次即可建立 E學中心.xlsx 範本"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(script_dir, "E學中心.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "帳號資料"

# 標題
headers = ["號碼", "姓名", "帳號", "密碼"]
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# 範例資料（可刪除）
ws.append(["1", "王小明", "my_egov_account", "my_egov_password"])

# 說明
ws.append([])
ws.append(["※ 帳號/密碼 = 我的e政府（cp.gov.tw）的帳號密碼"])

# 欄寬
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18

wb.save(out_path)
print(f"已建立：{out_path}")
print("請開啟 Excel 填入您的帳號密碼，完成後儲存。")
input("按 Enter 結束...")
