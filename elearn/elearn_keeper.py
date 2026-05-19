"""
E學中心 防登出 + 自動登入工具
使用 Selenium 開啟 Chrome，自動登入並持續保持 session 活躍
"""
import time
import sys
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, UnexpectedAlertPresentException
)

BASE_URL    = "https://elearn.hrd.gov.tw/mooc"
HOME_URL    = f"{BASE_URL}/index.php"
COURSE_URL  = f"{BASE_URL}/user/learn_dashboard.php?tab=1"

# ── 防登出腳本（注入到每個頁面） ──────────────────────────────────────────
KEEPALIVE_JS = """
if (!window._keepAliveRunning) {
    window._keepAliveRunning = true;
    var _tick = 0;

    // 攔截超過 5 分鐘的 setTimeout（登出計時器通常設這個範圍）
    var _origTimeout = window.setTimeout;
    window.setTimeout = function(fn, delay) {
        if (typeof delay === 'number' && delay > 5 * 60 * 1000) {
            console.log('[防登出] 攔截計時器 ' + Math.round(delay/1000) + 's -> 停用');
            return 0;
        }
        return _origTimeout.apply(this, arguments);
    };

    // 每 25 秒模擬活動
    setInterval(function() {
        _tick++;
        document.dispatchEvent(new MouseEvent('mousemove', {
            bubbles: true, clientX: 200 + (_tick % 50), clientY: 200 + (_tick % 30)
        }));
        window.scrollBy(0, 1);
        window.scrollBy(0, -1);
        console.log('[防登出] 保持活躍 #' + _tick);
    }, 25000);

    console.log('[防登出] 已啟動');
}
"""

# ────────────────────────────────────────────────────────────────────────────

def read_credentials(excel_path):
    """從 Excel 讀取帳號資料，回傳 list of dict"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Excel 欄位：{headers}")

    users = []
    for row in range(2, ws.max_row + 1):
        data = {headers[c-1]: ws.cell(row, c).value for c in range(1, len(headers)+1)}
        account = str(data.get('帳號') or data.get('account') or '').strip()
        passwd  = str(data.get('密碼') or data.get('password') or data.get('passwd') or '').strip()
        name    = str(data.get('姓名') or data.get('name') or '').strip()
        if account and passwd:
            users.append({'account': account, 'passwd': passwd, 'name': name})
    return users


def create_driver():
    """建立 Chrome，偽裝成正常瀏覽器"""
    opt = Options()
    opt.add_argument('--start-maximized')
    opt.add_argument('--disable-blink-features=AutomationControlled')
    opt.add_experimental_option('excludeSwitches', ['enable-automation'])
    opt.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=opt)
    driver.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
    return driver


def dismiss_popup(driver, timeout=4):
    """關閉「了解，我清楚了」等彈窗/通知按鈕"""
    xp = "//button[contains(.,'了解') or contains(.,'確定') or contains(.,'知道了') or contains(.,'關閉')]"
    try:
        btn = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xp)))
        btn.click()
        time.sleep(0.8)
        print("  ✓ 關閉通知彈窗")
    except TimeoutException:
        pass


def fill_field(driver, selectors, value, label):
    """嘗試多個 CSS selector，找到就填值，回傳成功/失敗"""
    for sel in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            if el.is_displayed():
                el.clear()
                el.send_keys(value)
                print(f"  ✓ 填入{label}（{sel}）")
                return True
        except NoSuchElementException:
            continue
    return False


def login(driver, account, passwd):
    """完整登入流程：E學中心 → 我的e政府 SSO → 成功回傳 True"""
    print(f"\n→ 登入帳號：{account}")

    # ── 步驟 1：直接跳到「我的e政府」SSO 登入頁（跳過 E學中心 UI） ────────
    SSO_URL = "https://www.cp.gov.tw/portal/portallogin.aspx?ReturnUrl=https://elearn.hrd.gov.tw/egov_login.php"
    driver.get(SSO_URL)
    time.sleep(3)
    print("  ✓ 到達 cp.gov.tw 登入選擇頁")

    # ── 步驟 8：點「我的E政府帳號登入」────────────────────────────────────
    try:
        acc_btn = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.ID, "accountlinkbt"))
        )
        acc_btn.click()
        print("  ✓ 點擊「我的E政府帳號登入」")
    except TimeoutException:
        print("  ✗ 找不到帳號登入按鈕")
        return False

    # ── 步驟 9：等待帳號密碼表單出現（postback 後頁面更新）─────────────
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "input[type='password'], input[type='text']:not([type='hidden'])"))
        )
        print("  ✓ 帳號密碼表單已出現")
        time.sleep(1)
    except TimeoutException:
        print(f"  ✗ 表單未出現（目前：{driver.current_url}）")
        # 印出所有 input 供除錯
        for i in driver.find_elements(By.TAG_NAME, "input"):
            print(f"     {i.get_attribute('id')} / {i.get_attribute('name')} / {i.get_attribute('type')}")
        return False

    # ── 步驟 10：填入帳號密碼 ────────────────────────────────────────────
    # 先找所有可見 text/password input
    all_inputs = driver.find_elements(By.XPATH,
        "//input[@type='text' or @type='password']")
    visible = [i for i in all_inputs if i.is_displayed()]
    print(f"  → 找到 {len(visible)} 個可見欄位")

    if len(visible) >= 2:
        visible[0].clear(); visible[0].send_keys(account)
        visible[1].clear(); visible[1].send_keys(passwd)
        print("  ✓ 填入帳號密碼")
    elif len(visible) == 1:
        visible[0].clear(); visible[0].send_keys(passwd)
        print("  ✓ 填入密碼（單欄）")
    else:
        print("  ✗ 找不到可見的帳號密碼欄位")
        return False

    # ── 步驟 11：點登入 ───────────────────────────────────────────────────
    submitted = False
    for sel in ["input[type='submit']", "button[type='submit']",
                "button.btn-login", "button.btn-primary",
                "button:not([type='button'])"]:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, sel)
            if btn.is_displayed():
                btn.click()
                submitted = True
                print(f"  ✓ 送出登入")
                break
        except Exception:
            continue

    if not submitted:
        visible[-1].send_keys(Keys.RETURN)
        print("  ✓ Enter 送出")

    # ── 步驟 12：等待跳回 E學中心，處理可能的 Alert ──────────────────────
    try:
        WebDriverWait(driver, 20).until(
            EC.url_contains("elearn.hrd.gov.tw")
        )
        print("  ✓ 已跳回 E學中心")
        time.sleep(2)
    except UnexpectedAlertPresentException:
        try:
            msg = driver.switch_to.alert.text
            driver.switch_to.alert.accept()
            print(f"  ✗ 登入錯誤：{msg[:80]}")
        except Exception:
            pass
        return False
    except TimeoutException:
        try:
            msg = driver.switch_to.alert.text
            driver.switch_to.alert.accept()
            print(f"  ✗ 登入錯誤：{msg[:80]}")
            return False
        except Exception:
            pass
        print(f"  ✗ 未跳回 E學中心（目前：{driver.current_url}）")
        return False

    # ── 步驟 13：確認登入成功 ─────────────────────────────────────────────
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "a[href*='logout'], a[href*='learn_dashboard'], .user-info"))
        )
        print("  ✓ 登入成功！")
        return True
    except TimeoutException:
        pass

    if "elearn.hrd.gov.tw" in driver.current_url and "login" not in driver.current_url:
        print("  ✓ 登入成功（URL 確認）")
        return True

    print("  ✗ 登入失敗")
    return False


def inject_keepalive(driver):
    """注入防登出腳本"""
    try:
        driver.execute_script(KEEPALIVE_JS)
    except Exception:
        pass  # 頁面轉換中，下次再注入


def go_to_courses(driver):
    """前往「未完成課程」"""
    driver.get(COURSE_URL)
    time.sleep(2)
    inject_keepalive(driver)
    print("  ✓ 已到「未完成課程」頁面")


def is_logged_in(driver):
    """快速確認是否還在登入狀態"""
    try:
        driver.find_element(By.CSS_SELECTOR,
            "a[href*='logout'], .user-info, a[href*='learn_dashboard']")
        return True
    except NoSuchElementException:
        return 'logout' not in driver.current_url and 'login' not in driver.current_url


def run(driver, account, passwd, check_interval=60):
    """主監控迴圈"""
    print(f"\n{'='*50}")
    print(" 監控啟動，按 Ctrl+C 可隨時停止")
    print(f" 每 {check_interval} 秒注入一次防登出腳本")
    print(f"{'='*50}\n")

    tick = 0
    while True:
        try:
            time.sleep(check_interval)
            tick += 1
            inject_keepalive(driver)

            # 每 10 次檢查一次登入狀態
            if tick % 10 == 0:
                if not is_logged_in(driver):
                    print(f"\n⚠️  [{tick}] 偵測到登出，重新登入...")
                    if login(driver, account, passwd):
                        go_to_courses(driver)
                else:
                    print(f"  [{tick}] ✓ 仍在登入中")
            else:
                print(f"  [{tick}] 保持活躍中...")

        except KeyboardInterrupt:
            print("\n\n已停止。")
            break
        except UnexpectedAlertPresentException:
            try:
                alert = driver.switch_to.alert
                alert.accept()
            except Exception:
                pass
        except Exception as e:
            print(f"  [{tick}] 警告：{e}")


# ── 主程式 ───────────────────────────────────────────────────────────────────

def main():
    print("=" * 50)
    print("   E學中心 防登出 + 自動登入工具  v1.0")
    print("=" * 50)

    # 找 Excel
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = None
    for name in ["E學中心.xlsx", "E學中心帳號.xlsx", "帳號.xlsx"]:
        candidate = os.path.join(script_dir, name)
        if os.path.exists(candidate):
            excel_path = candidate
            break

    if not excel_path:
        print("\n✗ 找不到 E學中心.xlsx！")
        print("  請確認 Excel 檔案和本程式放在同一個資料夾。")
        input("按 Enter 結束...")
        sys.exit(1)

    print(f"\n讀取：{os.path.basename(excel_path)}")
    users = read_credentials(excel_path)

    if not users:
        print("✗ Excel 中沒有找到帳號資料，請確認格式。")
        input("按 Enter 結束...")
        sys.exit(1)

    # 選擇帳號
    print(f"\n找到 {len(users)} 筆帳號：")
    for i, u in enumerate(users, 1):
        print(f"  {i}. {u['name']}  ({u['account']})")

    if len(users) > 1:
        choice = input(f"\n請輸入編號 (1~{len(users)})，直接 Enter 選第一筆：").strip()
        try:
            user = users[int(choice) - 1]
        except Exception:
            user = users[0]
    else:
        user = users[0]

    print(f"\n使用帳號：{user['name']} ({user['account']})")

    # 啟動瀏覽器
    print("\n啟動 Chrome...")
    driver = create_driver()

    try:
        if login(driver, user['account'], user['passwd']):
            go_to_courses(driver)
            run(driver, user['account'], user['passwd'])
        else:
            print("\n登入失敗，請檢查帳號密碼後重試。")
            input("按 Enter 關閉...")
    except Exception as e:
        print(f"\n錯誤：{e}")
        input("按 Enter 關閉...")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
