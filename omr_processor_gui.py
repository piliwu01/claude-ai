#!/usr/bin/env python3
"""
OMR 讀卡結果處理工具（圖形介面版）
雙擊此檔案即可啟動，不需要打開終端機。

必要套件（第一次使用前，在終端機執行）：
  pip install pandas openpyxl xlrd

若讀卡 .xls 檔無法開啟，需另安裝 LibreOffice：
  https://www.libreoffice.org/download/download-libreoffice/
"""

import os
import re
import sys
import shutil
import tempfile
import subprocess
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ── 核心處理邏輯（同命令列版）──────────────────────────────────

def _convert_with_libreoffice(xls_path, tmpdir):
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', tmpdir, xls_path],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice 轉換失敗：\n{result.stderr}")
    basename = os.path.splitext(os.path.basename(xls_path))[0] + '.xlsx'
    return os.path.join(tmpdir, basename)


def _find_header_row(df):
    keywords = ['班級', '座號', '姓名', '答案']
    for i, row in df.iterrows():
        text = ' '.join(str(v) for v in row.values if str(v) != 'nan')
        if sum(1 for kw in keywords if kw in text) >= 3:
            return i
    raise ValueError("找不到標題列（需含班級、座號、姓名、答案）")


def _find_columns(header_row):
    col_map = {}
    for idx, val in enumerate(header_row):
        txt = str(val)
        if ('班級' in txt or '座號' in txt) and '班級座號' not in col_map:
            col_map['班級座號'] = idx
        elif '姓名' in txt and '姓名' not in col_map:
            col_map['姓名'] = idx
        elif '答案' in txt and '答案' not in col_map:
            col_map['答案'] = idx
    return col_map


def _process(input_path, output_path, use_libreoffice, log):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    tmpdir = None
    xlsx_path = input_path

    # Step 1：轉檔
    if input_path.lower().endswith('.xls'):
        if use_libreoffice:
            log("🔄  使用 LibreOffice 轉換 .xls → .xlsx …")
            tmpdir = tempfile.mkdtemp()
            xlsx_path = _convert_with_libreoffice(input_path, tmpdir)
        else:
            try:
                pd.read_excel(input_path, header=None, nrows=1)
            except Exception:
                raise RuntimeError(
                    "無法直接讀取此 .xls 檔。\n\n"
                    "請勾選「使用 LibreOffice 轉換」後再試，\n"
                    "或先用 LibreOffice 手動另存為 .xlsx。"
                )

    # Step 2：解析
    log("📖  讀取並解析資料 …")
    df_raw = pd.read_excel(xlsx_path, header=None, sheet_name=0)
    header_idx = _find_header_row(df_raw)
    col_map = _find_columns(df_raw.iloc[header_idx])

    if len(col_map) < 3:
        raise ValueError(f"欄位偵測失敗，僅找到：{list(col_map.keys())}")

    ci_id   = col_map['班級座號']
    ci_name = col_map['姓名']
    ci_ans  = col_map['答案']

    data_rows, all_answers = [], []
    for i in range(header_idx + 1, len(df_raw)):
        val = df_raw.iloc[i, ci_id]
        if not re.match(r'^\d{4,6}$', str(val).strip()):
            continue
        raw = df_raw.iloc[i, ci_ans]
        ans = '' if pd.isna(raw) or str(raw).lower() == 'nan' else str(raw)
        data_rows.append({'班級座號': str(val).strip(), '姓名': df_raw.iloc[i, ci_name], '答案': ans})
        all_answers.append(ans)

    if not data_rows:
        raise ValueError("未找到學生資料，請確認檔案格式。")

    n_q = max((len(a) for a in all_answers), default=40) or 40
    mapping = {'A': 1, 'B': 2, 'C': 3, 'D': 4}
    special = sorted(set(''.join(all_answers)) - set('ABCD=. '))

    log(f"✅  找到 {len(data_rows)} 位學生，{n_q} 題")

    # Step 3：建立輸出 DataFrame
    log("🔢  轉換答案為數值 …")
    output_rows = []
    for row in data_rows:
        ans_str = row['答案'].ljust(n_q)[:n_q]
        out = {'班級座號': row['班級座號'], '姓名': row['姓名']}
        for j, ch in enumerate(ans_str):
            out[f'Q{j+1:02d}'] = mapping.get(ch, 0)
        output_rows.append(out)
    df_out = pd.DataFrame(output_rows)

    # Step 4：寫出 Excel
    log("💾  寫出 Excel …")
    wb = Workbook()
    ws = wb.active
    ws.title = '讀卡結果'

    hdr_fill   = PatternFill('solid', start_color='2F5496')
    hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    data_font  = Font(name='Arial', size=10)
    center     = Alignment(horizontal='center', vertical='center')
    thin       = __import__('openpyxl.styles', fromlist=['Side']).Side(style='thin', color='CCCCCC')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill   = PatternFill('solid', start_color='EEF2FF')
    white_fill = PatternFill('solid', start_color='FFFFFF')

    for ci, h in enumerate(df_out.columns, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, border

    for ri, row in df_out.iterrows():
        fill = alt_fill if ri % 2 == 1 else white_fill
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri + 2, column=ci, value=val)
            c.font, c.alignment, c.fill, c.border = data_font, center, fill, border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    for i in range(3, 3 + n_q):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 5
    ws.freeze_panes = 'A2'
    wb.save(output_path)

    if tmpdir:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return len(df_out), n_q, special


# ── GUI ────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('OMR 讀卡結果處理工具')
        self.resizable(False, False)
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f'+{(sw-w)//2}+{(sh-h)//2}')

    def _build_ui(self):
        PAD = {'padx': 14, 'pady': 6}

        # 標題
        tk.Label(self, text='📊 OMR 讀卡結果處理工具',
                 font=('Arial', 14, 'bold'), fg='#2F5496').pack(padx=14, pady=(14, 4))
        tk.Label(self, text='將讀卡機 XLS 轉換為每題一欄的數值化 Excel',
                 font=('Arial', 10), fg='#555').pack(pady=(0, 8))

        ttk.Separator(self).pack(fill='x', padx=14)

        # 輸入檔案
        frm_in = tk.Frame(self)
        frm_in.pack(fill='x', padx=14, pady=(10, 2))
        tk.Label(frm_in, text='輸入檔案：', width=10, anchor='w').pack(side='left')
        self.var_input = tk.StringVar()
        tk.Entry(frm_in, textvariable=self.var_input, width=46,
                 state='readonly', readonlybackground='white').pack(side='left', padx=(0, 6))
        ttk.Button(frm_in, text='選擇…', command=self._pick_input).pack(side='left')

        # 輸出檔案
        frm_out = tk.Frame(self)
        frm_out.pack(fill='x', padx=14, pady=2)
        tk.Label(frm_out, text='輸出檔案：', width=10, anchor='w').pack(side='left')
        self.var_output = tk.StringVar()
        tk.Entry(frm_out, textvariable=self.var_output, width=46,
                 state='readonly', readonlybackground='white').pack(side='left', padx=(0, 6))
        ttk.Button(frm_out, text='另存…', command=self._pick_output).pack(side='left')

        # LibreOffice 選項
        frm_opt = tk.Frame(self)
        frm_opt.pack(fill='x', padx=14, pady=(6, 2))
        self.var_lo = tk.BooleanVar(value=False)
        tk.Checkbutton(frm_opt, text='使用 LibreOffice 轉換（.xls 讀取失敗時勾選）',
                       variable=self.var_lo).pack(anchor='w')

        # 執行按鈕
        self.btn_run = ttk.Button(self, text='▶  開始處理', command=self._run)
        self.btn_run.pack(pady=(10, 4))

        ttk.Separator(self).pack(fill='x', padx=14)

        # 進度 / 訊息
        self.txt_log = tk.Text(self, height=8, width=60, state='disabled',
                               font=('Consolas', 10), bg='#f7f7f7',
                               relief='flat', padx=6, pady=4)
        self.txt_log.pack(padx=14, pady=(6, 14))

        # 開啟資料夾按鈕（初始隱藏）
        self.btn_open = ttk.Button(self, text='📂  開啟輸出資料夾', command=self._open_folder)
        self._output_dir = None

    def _pick_input(self):
        path = filedialog.askopenfilename(
            title='選擇讀卡 XLS / XLSX 檔案',
            filetypes=[('Excel 檔案', '*.xls *.xlsx'), ('所有檔案', '*.*')]
        )
        if path:
            self.var_input.set(path)
            base = os.path.splitext(path)[0]
            self.var_output.set(base + '_processed.xlsx')

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title='另存輸出 Excel',
            defaultextension='.xlsx',
            filetypes=[('Excel 檔案', '*.xlsx')]
        )
        if path:
            self.var_output.set(path)

    def _log(self, msg):
        self.txt_log.configure(state='normal')
        self.txt_log.insert('end', msg + '\n')
        self.txt_log.see('end')
        self.txt_log.configure(state='disabled')
        self.update_idletasks()

    def _run(self):
        inp = self.var_input.get()
        out = self.var_output.get()
        if not inp:
            messagebox.showwarning('提示', '請先選擇輸入檔案。')
            return
        if not out:
            messagebox.showwarning('提示', '請先設定輸出檔案路徑。')
            return

        self.btn_run.configure(state='disabled')
        self.btn_open.pack_forget()
        self.txt_log.configure(state='normal')
        self.txt_log.delete('1.0', 'end')
        self.txt_log.configure(state='disabled')

        threading.Thread(target=self._worker, args=(inp, out), daemon=True).start()

    def _worker(self, inp, out):
        try:
            n_s, n_q, special = _process(inp, out, self.var_lo.get(), self._log)
            msg = f'\n🎉  完成！\n    {n_s} 位學生 × {n_q} 題'
            if special:
                msg += f'\n    ⚠  非標準字元（複選碼）已轉為 0：{special}'
            self._log(msg)
            self._output_dir = os.path.dirname(os.path.abspath(out))
            self.btn_open.pack(pady=(0, 14))
        except Exception as e:
            self._log(f'\n❌  錯誤：{e}')
        finally:
            self.btn_run.configure(state='normal')

    def _open_folder(self):
        if self._output_dir and os.path.isdir(self._output_dir):
            if sys.platform == 'win32':
                os.startfile(self._output_dir)
            elif sys.platform == 'darwin':
                subprocess.run(['open', self._output_dir])
            else:
                subprocess.run(['xdg-open', self._output_dir])


if __name__ == '__main__':
    App().mainloop()
