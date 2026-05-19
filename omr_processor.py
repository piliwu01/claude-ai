#!/usr/bin/env python3
"""
OMR 讀卡結果處理工具
================================
將讀卡機產出的 XLS/XLSX 檔案轉換為每生一列、每題一欄的數值化 Excel。

使用方式：
  python omr_processor.py 讀卡.xls
  python omr_processor.py 讀卡.xls -o 結果.xlsx

必要套件（第一次使用前執行）：
  pip install pandas openpyxl xlrd

若輸入為 .xls 且 xlrd 無法讀取，需安裝 LibreOffice 並加上 --libreoffice 旗標：
  python omr_processor.py 讀卡.xls --libreoffice
"""

import sys
import os
import re
import argparse
import subprocess
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── 核心處理函式 ──────────────────────────────────────────────

def convert_with_libreoffice(xls_path: str, tmpdir: str) -> str:
    """用 LibreOffice 將 .xls 轉為 .xlsx，回傳新路徑。"""
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', tmpdir, xls_path],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice 轉換失敗：\n{result.stderr}")
    basename = os.path.splitext(os.path.basename(xls_path))[0] + '.xlsx'
    return os.path.join(tmpdir, basename)


def find_header_row(df_raw: pd.DataFrame) -> int:
    """找出含有班級/座號/姓名/答案關鍵字的標題列索引。"""
    keywords = ['班級', '座號', '姓名', '答案']
    for i, row in df_raw.iterrows():
        text = ' '.join(str(v) for v in row.values if pd.notna(v))
        if sum(1 for kw in keywords if kw in text) >= 3:
            return i
    raise ValueError("找不到標題列（需含班級、座號、姓名、答案等關鍵字），請確認檔案格式。")


def find_columns(header_row: pd.Series) -> dict:
    """從標題列找出各欄位的索引。"""
    col_map = {}
    for idx, val in enumerate(header_row):
        txt = str(val)
        if ('班級' in txt or '座號' in txt) and '班級座號' not in col_map:
            col_map['班級座號'] = idx
        elif '姓名' in txt and '姓名' not in col_map:
            col_map['姓名'] = idx
        elif '答案' in txt and '答案' not in col_map:
            col_map['答案'] = idx
    missing = [k for k in ['班級座號', '姓名', '答案'] if k not in col_map]
    if missing:
        raise ValueError(f"缺少欄位：{missing}，偵測到的欄位：{col_map}")
    return col_map


def is_student_row(val) -> bool:
    """判斷是否為有效學生列（班級座號為 4～6 位數字）。"""
    return bool(re.match(r'^\d{4,6}$', str(val).strip()))


def split_answers(ans_str: str, n_questions: int) -> list:
    """切分答案字串為數值 list，A=1 B=2 C=3 D=4，其餘=0。"""
    mapping = {'A': 1, 'B': 2, 'C': 3, 'D': 4}
    ans_str = str(ans_str).ljust(n_questions)[:n_questions]
    return [mapping.get(ch, 0) for ch in ans_str]


def write_excel(df_out: pd.DataFrame, output_path: str, n_questions: int):
    """將處理後的 DataFrame 寫出為格式化 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = '讀卡結果'

    header_fill = PatternFill('solid', start_color='2F5496')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    data_font   = Font(name='Arial', size=10)
    center      = Alignment(horizontal='center', vertical='center')
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill    = PatternFill('solid', start_color='EEF2FF')
    white_fill  = PatternFill('solid', start_color='FFFFFF')

    # 標題列
    for col_idx, h in enumerate(df_out.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border

    # 資料列
    for row_idx, row in df_out.iterrows():
        fill = alt_fill if row_idx % 2 == 1 else white_fill
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.font, cell.alignment, cell.fill, cell.border = data_font, center, fill, border

    # 欄寬
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    for i in range(3, 3 + n_questions):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 5

    ws.freeze_panes = 'A2'
    wb.save(output_path)


def process(input_path: str, output_path: str, use_libreoffice: bool = False):
    """主流程：讀取 → 清理 → 切分 → 寫出。"""
    # Step 1：轉檔
    xlsx_path = input_path
    tmpdir = None
    if input_path.lower().endswith('.xls'):
        if use_libreoffice:
            tmpdir = tempfile.mkdtemp()
            print(f"[1/3] 使用 LibreOffice 轉換 {os.path.basename(input_path)} ...")
            xlsx_path = convert_with_libreoffice(input_path, tmpdir)
        else:
            # 嘗試直接以 xlrd 讀取（部分 .xls 可行）
            try:
                pd.read_excel(input_path, header=None, nrows=1)
            except Exception:
                print("⚠  無法直接讀取 .xls，請加上 --libreoffice 旗標（需安裝 LibreOffice）")
                sys.exit(1)
            xlsx_path = input_path

    # Step 2：讀取並解析
    print(f"[2/3] 讀取並解析 {os.path.basename(xlsx_path)} ...")
    df_raw = pd.read_excel(xlsx_path, header=None, sheet_name=0)

    header_row_idx = find_header_row(df_raw)
    col_map = find_columns(df_raw.iloc[header_row_idx])
    ci_id, ci_name, ci_ans = col_map['班級座號'], col_map['姓名'], col_map['答案']

    data_rows, all_answers = [], []
    for i in range(header_row_idx + 1, len(df_raw)):
        val = df_raw.iloc[i, ci_id]
        if not is_student_row(val):
            continue
        raw_ans = df_raw.iloc[i, ci_ans]
        ans = '' if pd.isna(raw_ans) or str(raw_ans).lower() == 'nan' else str(raw_ans)
        data_rows.append({'班級座號': str(val).strip(), '姓名': df_raw.iloc[i, ci_name], '答案': ans})
        all_answers.append(ans)

    if not data_rows:
        raise ValueError("未找到學生資料列，請確認格式。")

    n_questions = max((len(a) for a in all_answers), default=40) or 40
    special_chars = sorted(set(''.join(all_answers)) - set('ABCD=. '))

    output_rows = []
    for row in data_rows:
        q_vals = split_answers(row['答案'], n_questions)
        out = {'班級座號': row['班級座號'], '姓名': row['姓名']}
        for j, v in enumerate(q_vals):
            out[f'Q{j+1:02d}'] = v
        output_rows.append(out)

    df_out = pd.DataFrame(output_rows)

    # Step 3：寫出
    print(f"[3/3] 寫出 {output_path} ...")
    write_excel(df_out, output_path, n_questions)

    # 清理暫存
    if tmpdir:
        import shutil
        shutil.rmtree(tmpdir, ignore_errors=True)

    return len(df_out), n_questions, special_chars


# ── 主程式進入點 ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='OMR 讀卡結果處理工具：XLS → 每題一欄的數值化 Excel'
    )
    parser.add_argument('input', help='輸入檔案（.xls 或 .xlsx）')
    parser.add_argument('-o', '--output', help='輸出檔案路徑（預設：原檔名_processed.xlsx）')
    parser.add_argument('--libreoffice', action='store_true',
                        help='使用 LibreOffice 轉換 .xls（遇到加密或舊格式時使用）')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"❌ 找不到檔案：{args.input}")
        sys.exit(1)

    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(args.input)[0]
        output_path = base + '_processed.xlsx'

    try:
        n_students, n_questions, special_chars = process(args.input, output_path, args.libreoffice)
        print(f"\n✅ 完成！")
        print(f"   學生人數：{n_students} 人")
        print(f"   題目數量：{n_questions} 題")
        print(f"   輸出路徑：{output_path}")
        if special_chars:
            print(f"   ⚠  以下非標準字元已轉為 0（複選或特殊碼）：{special_chars}")
    except Exception as e:
        print(f"\n❌ 處理失敗：{e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
